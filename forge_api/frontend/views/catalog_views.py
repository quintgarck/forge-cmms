"""
Views for expanded catalog management interfaces.
Handles EquipmentType, TaxonomySystem, TaxonomySubsystem, TaxonomyGroup,
reference codes (Fuel, Transmission, Color, etc.), Currency, and Supplier management.
"""
import logging
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView
from django.views import View
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.conf import settings

logger = logging.getLogger(__name__)

from ..services.api_client import ForgeAPIClient, APIException
from ..mixins import APIClientMixin

logger = logging.getLogger(__name__)


class CatalogIndexView(LoginRequiredMixin, TemplateView):
    """Vista principal del índice de catálogos"""
    template_name = 'frontend/catalog/catalog_index.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Catálogos'
        return context


class EquipmentTypeListView(LoginRequiredMixin, APIClientMixin, TemplateView):
    """Equipment type list view with hierarchical display."""
    template_name = 'frontend/catalog/equipment_type_list.html'
    login_url = 'frontend:login'
    paginate_by = 20
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get search and filter parameters
        search = self.request.GET.get('search', '').strip()
        page = self._get_page_number()
        category_filter = self.request.GET.get('category', '')
        sort_by = self.request.GET.get('sort', 'name')
        sort_order = self.request.GET.get('order', 'asc')
        
        try:
            api_client = self.get_api_client()
            
            # Build filter parameters
            filters = {}
            if search:
                filters['search'] = search
            if category_filter:
                filters['category'] = category_filter
            if sort_by:
                order_prefix = '-' if sort_order == 'desc' else ''
                filters['ordering'] = f"{order_prefix}{sort_by}"
            
            # Get equipment types data
            equipment_types_data = api_client.get('equipment-types/', params={
                'page': page,
                'page_size': self.paginate_by,
                **filters
            })
            
            context['equipment_types'] = equipment_types_data.get('results', [])
            
            # Process equipment types for hierarchical display
            for eq_type in context['equipment_types']:
                # Add category styling
                category = eq_type.get('category', '').lower()
                eq_type['category_class'] = self._get_category_class(category)
                eq_type['category_icon'] = self._get_category_icon(category)
                
                # Process attributes for display
                attributes = eq_type.get('attr_schema', {})
                if isinstance(attributes, dict):
                    eq_type['attribute_count'] = len(attributes)
                    eq_type['key_attributes'] = list(attributes.keys())[:3]  # Show first 3
                else:
                    eq_type['attribute_count'] = 0
                    eq_type['key_attributes'] = []
            
            # Enhanced pagination context
            total_count = equipment_types_data.get('count', 0)
            total_pages = (total_count + self.paginate_by - 1) // self.paginate_by
            
            context['pagination'] = {
                'count': total_count,
                'next': equipment_types_data.get('next'),
                'previous': equipment_types_data.get('previous'),
                'current_page': page,
                'total_pages': total_pages,
                'page_range': self._get_page_range(page, total_pages),
                'has_previous': page > 1,
                'has_next': page < total_pages,
                'start_index': (page - 1) * self.paginate_by + 1,
                'end_index': min(page * self.paginate_by, total_count),
            }
            
            # Filter context
            context['filters'] = {
                'search': search,
                'category': category_filter,
                'sort': sort_by,
                'order': sort_order,
            }
            
            # Category options for filter
            context['category_options'] = [
                {'value': '', 'label': 'Todas las Categorías'},
                {'value': 'automotive', 'label': 'Automotriz'},
                {'value': 'heavy_machinery', 'label': 'Maquinaria Pesada'},
                {'value': 'marine', 'label': 'Marino'},
                {'value': 'industrial', 'label': 'Industrial'},
                {'value': 'agricultural', 'label': 'Agrícola'},
            ]
            
        except APIException as e:
            self.handle_api_error(e, "Error al cargar tipos de equipo")
            context['equipment_types'] = []
            context['pagination'] = self._get_empty_pagination()
            context['filters'] = {
                'search': search,
                'category': category_filter,
                'sort': sort_by,
                'order': sort_order,
            }
        
        return context
    
    def _get_category_class(self, category):
        """Get Bootstrap class for equipment category."""
        category_classes = {
            'automotive': 'primary',
            'heavy_machinery': 'warning',
            'marine': 'info',
            'industrial': 'success',
            'agricultural': 'secondary',
        }
        return category_classes.get(category, 'light')
    
    def _get_category_icon(self, category):
        """Get icon for equipment category."""
        category_icons = {
            'automotive': 'bi-car-front',
            'heavy_machinery': 'bi-truck',
            'marine': 'bi-water',
            'industrial': 'bi-gear-wide-connected',
            'agricultural': 'bi-tree',
        }
        return category_icons.get(category, 'bi-gear')
    
    def _get_page_number(self):
        """Get and validate page number from request."""
        try:
            page = int(self.request.GET.get('page', 1))
            return max(1, page)
        except (ValueError, TypeError):
            return 1
    
    def _get_page_range(self, current_page, total_pages, window=5):
        """Generate a smart page range for pagination."""
        if total_pages <= window:
            return list(range(1, total_pages + 1))
        
        half_window = window // 2
        start = max(1, current_page - half_window)
        end = min(total_pages, current_page + half_window)
        
        if end - start < window - 1:
            if start == 1:
                end = min(total_pages, start + window - 1)
            else:
                start = max(1, end - window + 1)
        
        return list(range(start, end + 1))
    
    def _get_empty_pagination(self):
        """Get empty pagination context for error states."""
        return {
            'count': 0,
            'current_page': 1,
            'total_pages': 0,
            'page_range': [],
            'has_previous': False,
            'has_next': False,
            'start_index': 0,
            'end_index': 0,
        }


class EquipmentTypeDetailView(LoginRequiredMixin, APIClientMixin, TemplateView):
    """Equipment type detail view with attributes and relationships."""
    template_name = 'frontend/catalog/equipment_type_detail.html'
    login_url = 'frontend:login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        equipment_type_id = kwargs.get('pk')
        
        try:
            api_client = self.get_api_client()
            
            # Get equipment type details
            equipment_type = api_client.get(f'equipment-types/{equipment_type_id}/')
            context['equipment_type'] = equipment_type
            
            # Process attributes for better display
            attributes = equipment_type.get('attributes', {})
            if isinstance(attributes, dict):
                context['attributes'] = [
                    {'key': key, 'value': value}
                    for key, value in attributes.items()
                ]
            else:
                context['attributes'] = []
            
            # Get related equipment count
            try:
                related_equipment = api_client.get('equipment/', params={
                    'equipment_type': equipment_type_id,
                    'page_size': 1
                })
                context['related_equipment_count'] = related_equipment.get('count', 0)
            except APIException:
                context['related_equipment_count'] = 0
            
            # Get taxonomy information if available
            if equipment_type.get('taxonomy_group_id'):
                try:
                    taxonomy_group = api_client.get(f'taxonomy-groups/{equipment_type["taxonomy_group_id"]}/')
                    context['taxonomy_group'] = taxonomy_group
                    
                    # Get parent taxonomy information
                    if taxonomy_group.get('taxonomy_subsystem_id'):
                        taxonomy_subsystem = api_client.get(f'taxonomy-subsystems/{taxonomy_group["taxonomy_subsystem_id"]}/')
                        context['taxonomy_subsystem'] = taxonomy_subsystem
                        
                        if taxonomy_subsystem.get('taxonomy_system_id'):
                            taxonomy_system = api_client.get(f'taxonomy-systems/{taxonomy_subsystem["taxonomy_system_id"]}/')
                            context['taxonomy_system'] = taxonomy_system
                            
                except APIException:
                    pass  # Taxonomy information is optional
            
        except APIException as e:
            self.handle_api_error(e, "Error al cargar detalles del tipo de equipo")
            context['equipment_type'] = None
        
        return context


class TaxonomySystemListView(LoginRequiredMixin, APIClientMixin, TemplateView):
    """Taxonomy system list view with hierarchical structure."""
    template_name = 'frontend/catalog/taxonomy_system_list.html'
    login_url = 'frontend:login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        try:
            api_client = self.get_api_client()
            
            # Get taxonomy systems
            taxonomy_systems = api_client.get('taxonomy-systems/')
            context['taxonomy_systems'] = taxonomy_systems.get('results', [])
            
            # For each system, get subsystems and groups for hierarchical display
            for system in context['taxonomy_systems']:
                try:
                    # Get subsystems for this system
                    subsystems = api_client.get('taxonomy-subsystems/', params={
                        'taxonomy_system': system['system_code']
                    })
                    system['subsystems'] = subsystems.get('results', [])
                    
                    # For each subsystem, get groups
                    for subsystem in system['subsystems']:
                        try:
                            groups = api_client.get('taxonomy-groups/', params={
                                'taxonomy_subsystem': subsystem['subsystem_code']
                            })
                            subsystem['groups'] = groups.get('results', [])
                        except APIException:
                            subsystem['groups'] = []
                            
                except APIException:
                    system['subsystems'] = []
            
        except APIException as e:
            self.handle_api_error(e, "Error al cargar sistemas de taxonomía")
            context['taxonomy_systems'] = []
        
        return context


class ReferenceCodeListView(LoginRequiredMixin, APIClientMixin, TemplateView):
    """Reference codes list view (Fuel, Transmission, Color, etc.)."""
    template_name = 'frontend/catalog/reference_code_list.html'
    login_url = 'frontend:login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        try:
            api_client = self.get_api_client()
            
            # Get all reference code types
            reference_codes = {}
            
            # Fuel codes
            try:
                fuel_codes = api_client.get('fuel-codes/')
                reference_codes['fuel'] = {
                    'title': 'Códigos de Combustible',
                    'icon': 'bi-fuel-pump',
                    'color': 'danger',
                    'codes': fuel_codes.get('results', [])
                }
            except APIException:
                reference_codes['fuel'] = {
                    'title': 'Códigos de Combustible',
                    'icon': 'bi-fuel-pump',
                    'color': 'danger',
                    'codes': []
                }
            
            # Transmission codes
            try:
                transmission_codes = api_client.get('transmission-codes/')
                reference_codes['transmission'] = {
                    'title': 'Códigos de Transmisión',
                    'icon': 'bi-gear-wide-connected',
                    'color': 'primary',
                    'codes': transmission_codes.get('results', [])
                }
            except APIException:
                reference_codes['transmission'] = {
                    'title': 'Códigos de Transmisión',
                    'icon': 'bi-gear-wide-connected',
                    'color': 'primary',
                    'codes': []
                }
            
            # Color codes
            try:
                color_codes = api_client.get('color-codes/')
                reference_codes['color'] = {
                    'title': 'Códigos de Color',
                    'icon': 'bi-palette',
                    'color': 'info',
                    'codes': color_codes.get('results', [])
                }
            except APIException:
                reference_codes['color'] = {
                    'title': 'Códigos de Color',
                    'icon': 'bi-palette',
                    'color': 'info',
                    'codes': []
                }
            
            # Drivetrain codes
            try:
                drivetrain_codes = api_client.get('drivetrain-codes/')
                reference_codes['drivetrain'] = {
                    'title': 'Códigos de Tracción',
                    'icon': 'bi-arrow-left-right',
                    'color': 'warning',
                    'codes': drivetrain_codes.get('results', [])
                }
            except APIException:
                reference_codes['drivetrain'] = {
                    'title': 'Códigos de Tracción',
                    'icon': 'bi-arrow-left-right',
                    'color': 'warning',
                    'codes': []
                }
            
            # Condition codes
            try:
                condition_codes = api_client.get('condition-codes/')
                reference_codes['condition'] = {
                    'title': 'Códigos de Condición',
                    'icon': 'bi-check-circle',
                    'color': 'success',
                    'codes': condition_codes.get('results', [])
                }
            except APIException:
                reference_codes['condition'] = {
                    'title': 'Códigos de Condición',
                    'icon': 'bi-check-circle',
                    'color': 'success',
                    'codes': []
                }
            
            # Aspiration codes
            try:
                aspiration_codes = api_client.get('aspiration-codes/')
                reference_codes['aspiration'] = {
                    'title': 'Códigos de Aspiración',
                    'icon': 'bi-wind',
                    'color': 'secondary',
                    'codes': aspiration_codes.get('results', [])
                }
            except APIException:
                reference_codes['aspiration'] = {
                    'title': 'Códigos de Aspiración',
                    'icon': 'bi-wind',
                    'color': 'secondary',
                    'codes': []
                }
            
            context['reference_codes'] = reference_codes
            
            # Calculate totals
            total_codes = sum(len(category['codes']) for category in reference_codes.values())
            context['total_codes'] = total_codes
            
        except APIException as e:
            self.handle_api_error(e, "Error al cargar códigos de referencia")
            context['reference_codes'] = {}
            context['total_codes'] = 0
        
        return context


class CurrencyListView(LoginRequiredMixin, APIClientMixin, TemplateView):
    """Currency management list view."""
    template_name = 'frontend/catalog/currency_list.html'
    login_url = 'frontend:login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        try:
            api_client = self.get_api_client()
            
            # Get currencies using the API client method
            currencies_data = api_client.get_currencies()
            currencies = currencies_data.get('results', [])
            
            # Process currencies for display
            for currency in currencies:
                # Add status styling
                if currency.get('is_active'):
                    currency['status_class'] = 'success'
                    currency['status_label'] = 'Activa'
                else:
                    currency['status_class'] = 'secondary'
                    currency['status_label'] = 'Inactiva'
                
                # Format exchange rate (convert to float if string)
                exchange_rate_raw = currency.get('exchange_rate', '1.0')
                try:
                    exchange_rate = float(exchange_rate_raw)
                except (ValueError, TypeError):
                    exchange_rate = 1.0
                currency['exchange_rate_formatted'] = f"{exchange_rate:.4f}"
                
                # Determine if is base currency (exchange_rate == 1.0)
                currency['is_base_currency'] = (exchange_rate == 1.0)
            
            context['currencies'] = currencies
            
            # Get base currency
            base_currency = next((c for c in currencies if c.get('is_base_currency')), None)
            context['base_currency'] = base_currency
            
        except APIException as e:
            self.handle_api_error(e, "Error al cargar monedas")
            context['currencies'] = []
            context['base_currency'] = None
        
        return context


class SupplierAdvancedListView(LoginRequiredMixin, APIClientMixin, TemplateView):
    """Advanced supplier management list view."""
    template_name = 'frontend/catalog/supplier_advanced_list.html'
    login_url = 'frontend:login'
    paginate_by = 20
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get search and filter parameters
        search = self.request.GET.get('search', '').strip()
        page = self._get_page_number()
        category_filter = self.request.GET.get('category', '')
        status_filter = self.request.GET.get('status', '')
        sort_by = self.request.GET.get('sort', 'name')
        sort_order = self.request.GET.get('order', 'asc')
        
        try:
            api_client = self.get_api_client()
            
            # Build filter parameters
            filters = {}
            if search:
                filters['search'] = search
            if category_filter:
                filters['category'] = category_filter
            if status_filter:
                filters['status'] = status_filter
            if sort_by:
                order_prefix = '-' if sort_order == 'desc' else ''
                filters['ordering'] = f"{order_prefix}{sort_by}"
            
            # Get suppliers data
            suppliers_data = api_client.get('suppliers/', params={
                'page': page,
                'page_size': self.paginate_by,
                **filters
            })
            
            suppliers = suppliers_data.get('results', [])
            
            # Process suppliers for enhanced display
            for supplier in suppliers:
                # Add status styling
                if supplier.get('is_active'):
                    supplier['status_class'] = 'success'
                    supplier['status_label'] = 'Activo'
                else:
                    supplier['status_class'] = 'secondary'
                    supplier['status_label'] = 'Inactivo'
                
                # Add category styling
                category = supplier.get('category', '').lower()
                supplier['category_class'] = self._get_supplier_category_class(category)
                supplier['category_icon'] = self._get_supplier_category_icon(category)
                
                # Calculate performance metrics (if available)
                supplier['performance_score'] = supplier.get('performance_score', 0)
                if supplier['performance_score'] >= 90:
                    supplier['performance_class'] = 'success'
                elif supplier['performance_score'] >= 70:
                    supplier['performance_class'] = 'warning'
                else:
                    supplier['performance_class'] = 'danger'
            
            context['suppliers'] = suppliers
            
            # Enhanced pagination context
            total_count = suppliers_data.get('count', 0)
            total_pages = (total_count + self.paginate_by - 1) // self.paginate_by
            
            context['pagination'] = {
                'count': total_count,
                'next': suppliers_data.get('next'),
                'previous': suppliers_data.get('previous'),
                'current_page': page,
                'total_pages': total_pages,
                'page_range': self._get_page_range(page, total_pages),
                'has_previous': page > 1,
                'has_next': page < total_pages,
                'start_index': (page - 1) * self.paginate_by + 1,
                'end_index': min(page * self.paginate_by, total_count),
            }
            
            # Filter context
            context['filters'] = {
                'search': search,
                'category': category_filter,
                'status': status_filter,
                'sort': sort_by,
                'order': sort_order,
            }
            
            # Filter options
            context['category_options'] = [
                {'value': '', 'label': 'Todas las Categorías'},
                {'value': 'parts', 'label': 'Repuestos'},
                {'value': 'services', 'label': 'Servicios'},
                {'value': 'materials', 'label': 'Materiales'},
                {'value': 'tools', 'label': 'Herramientas'},
                {'value': 'oem', 'label': 'OEM'},
            ]
            
            context['status_options'] = [
                {'value': '', 'label': 'Todos los Estados'},
                {'value': 'active', 'label': 'Activos'},
                {'value': 'inactive', 'label': 'Inactivos'},
                {'value': 'pending', 'label': 'Pendientes'},
            ]
            
        except APIException as e:
            self.handle_api_error(e, "Error al cargar proveedores")
            context['suppliers'] = []
            context['pagination'] = self._get_empty_pagination()
            context['filters'] = {
                'search': search,
                'category': category_filter,
                'status': status_filter,
                'sort': sort_by,
                'order': sort_order,
            }
        
        return context
    
    def _get_supplier_category_class(self, category):
        """Get Bootstrap class for supplier category."""
        category_classes = {
            'parts': 'primary',
            'services': 'success',
            'materials': 'info',
            'tools': 'warning',
            'oem': 'danger',
        }
        return category_classes.get(category, 'secondary')
    
    def _get_supplier_category_icon(self, category):
        """Get icon for supplier category."""
        category_icons = {
            'parts': 'bi-gear',
            'services': 'bi-tools',
            'materials': 'bi-box',
            'tools': 'bi-hammer',
            'oem': 'bi-building',
        }
        return category_icons.get(category, 'bi-shop')
    
    def _get_page_number(self):
        """Get and validate page number from request."""
        try:
            page = int(self.request.GET.get('page', 1))
            return max(1, page)
        except (ValueError, TypeError):
            return 1
    
    def _get_page_range(self, current_page, total_pages, window=5):
        """Generate a smart page range for pagination."""
        if total_pages <= window:
            return list(range(1, total_pages + 1))
        
        half_window = window // 2
        start = max(1, current_page - half_window)
        end = min(total_pages, current_page + half_window)
        
        if end - start < window - 1:
            if start == 1:
                end = min(total_pages, start + window - 1)
            else:
                start = max(1, end - window + 1)
        
        return list(range(start, end + 1))
    
    def _get_empty_pagination(self):
        """Get empty pagination context for error states."""
        return {
            'count': 0,
            'current_page': 1,
            'total_pages': 0,
            'page_range': [],
            'has_previous': False,
            'has_next': False,
            'start_index': 0,
            'end_index': 0,
        }



class CatalogReportsView(LoginRequiredMixin, APIClientMixin, TemplateView):
    """
    Vista de reportes y análisis de catálogo.
    Proporciona estadísticas, gráficos y análisis de todos los módulos del catálogo.
    Incluye filtros por fecha, tendencias y análisis predictivo.
    """
    template_name = 'frontend/catalog/catalog_reports.html'
    login_url = 'frontend:login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener filtros de fecha
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')
        period = self.request.GET.get('period', '30')  # días
        
        # Si no hay fechas específicas, usar período
        if not date_from or not date_to:
            try:
                days = int(period)
                date_to = timezone.now()
                date_from = date_to - timedelta(days=days)
            except ValueError:
                days = 30
                date_to = timezone.now()
                date_from = date_to - timedelta(days=days)
        else:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d')
                date_to = datetime.strptime(date_to, '%Y-%m-%d')
            except ValueError:
                date_to = timezone.now()
                date_from = date_to - timedelta(days=30)
        
        context['filters'] = {
            'date_from': date_from.strftime('%Y-%m-%d') if hasattr(date_from, 'strftime') else date_from,
            'date_to': date_to.strftime('%Y-%m-%d') if hasattr(date_to, 'strftime') else date_to,
            'period': period
        }
        
        try:
            api_client = self.get_api_client()
            
            # Estadísticas de Tipos de Equipo
            try:
                equipment_types_data = api_client.get('equipment-types/', params={'page_size': 1})
                context['equipment_types_count'] = equipment_types_data.get('count', 0)
            except APIException:
                context['equipment_types_count'] = 0
            
            # Estadísticas de Taxonomía
            try:
                taxonomy_systems = api_client.get('taxonomy-systems/', params={'page_size': 1})
                context['taxonomy_systems_count'] = taxonomy_systems.get('count', 0)
                
                taxonomy_subsystems = api_client.get('taxonomy-subsystems/', params={'page_size': 1})
                context['taxonomy_subsystems_count'] = taxonomy_subsystems.get('count', 0)
                
                taxonomy_groups = api_client.get('taxonomy-groups/', params={'page_size': 1})
                context['taxonomy_groups_count'] = taxonomy_groups.get('count', 0)
            except APIException:
                context['taxonomy_systems_count'] = 0
                context['taxonomy_subsystems_count'] = 0
                context['taxonomy_groups_count'] = 0
            
            # Estadísticas de Códigos de Referencia
            reference_code_counts = {}
            for category in ['fuel', 'transmission', 'color', 'drivetrain', 'condition', 'aspiration']:
                try:
                    codes_data = api_client.get(f'{category}-codes/', params={'page_size': 1})
                    reference_code_counts[category] = codes_data.get('count', 0)
                except APIException:
                    reference_code_counts[category] = 0
            
            context['reference_code_counts'] = reference_code_counts
            context['total_reference_codes'] = sum(reference_code_counts.values())
            
            # Estadísticas de Monedas
            try:
                currencies_data = api_client.get('currencies/', params={'page_size': 1})
                context['currencies_count'] = currencies_data.get('count', 0)
            except APIException:
                context['currencies_count'] = 0
            
            # Estadísticas de Proveedores
            try:
                suppliers_data = api_client.get('suppliers/', params={'page_size': 1})
                context['suppliers_count'] = suppliers_data.get('count', 0)
                
                # Proveedores activos
                active_suppliers = api_client.get('suppliers/', params={
                    'page_size': 1,
                    'is_active': 'true'
                })
                context['active_suppliers_count'] = active_suppliers.get('count', 0)
            except APIException:
                context['suppliers_count'] = 0
                context['active_suppliers_count'] = 0
            
            # Resumen general
            context['total_catalog_items'] = (
                context['equipment_types_count'] +
                context['taxonomy_systems_count'] +
                context['taxonomy_subsystems_count'] +
                context['taxonomy_groups_count'] +
                context['total_reference_codes'] +
                context['currencies_count'] +
                context['suppliers_count']
            )
            
            # Datos para gráficos (formato JSON para Chart.js)
            context['chart_data'] = {
                'reference_codes': {
                    'labels': ['Combustible', 'Transmisión', 'Color', 'Tracción', 'Condición', 'Aspiración'],
                    'data': [
                        reference_code_counts.get('fuel', 0),
                        reference_code_counts.get('transmission', 0),
                        reference_code_counts.get('color', 0),
                        reference_code_counts.get('drivetrain', 0),
                        reference_code_counts.get('condition', 0),
                        reference_code_counts.get('aspiration', 0),
                    ]
                },
                'taxonomy': {
                    'labels': ['Sistemas', 'Subsistemas', 'Grupos'],
                    'data': [
                        context['taxonomy_systems_count'],
                        context['taxonomy_subsystems_count'],
                        context['taxonomy_groups_count'],
                    ]
                }
            }
            
        except APIException as e:
            self.handle_api_error(e, "Error al cargar reportes de catálogo")
            # Valores por defecto en caso de error
            context['equipment_types_count'] = 0
            context['taxonomy_systems_count'] = 0
            context['taxonomy_subsystems_count'] = 0
            context['taxonomy_groups_count'] = 0
            context['reference_code_counts'] = {}
            context['total_reference_codes'] = 0
            context['currencies_count'] = 0
            context['suppliers_count'] = 0
            context['active_suppliers_count'] = 0
            context['total_catalog_items'] = 0
            context['chart_data'] = {
                'reference_codes': {'labels': [], 'data': []},
                'taxonomy': {'labels': [], 'data': []}
            }
        
        return context



class CatalogReportExportView(LoginRequiredMixin, APIClientMixin, View):
    """
    Vista para exportar reportes de catálogo en diferentes formatos.
    Soporta PDF y Excel.
    """
    login_url = 'frontend:login'
    
    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('format', 'pdf').lower()
        
        if export_format == 'pdf':
            return self._export_pdf(request)
        elif export_format == 'excel':
            return self._export_excel(request)
        else:
            messages.error(request, "Formato de exportación no válido")
            return redirect('frontend:catalog_reports')
    
    def _export_pdf(self, request):
        """Exportar reporte a PDF"""
        from django.http import HttpResponse
        from django.template.loader import render_to_string
        try:
            from weasyprint import HTML
            import tempfile
            
            # Obtener datos del reporte
            context = self._get_report_data()
            context['export_date'] = timezone.now()
            
            # Renderizar template HTML
            html_string = render_to_string('frontend/catalog/catalog_report_pdf.html', context)
            
            # Generar PDF
            html = HTML(string=html_string)
            pdf_file = html.write_pdf()
            
            # Crear respuesta HTTP
            response = HttpResponse(pdf_file, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="reporte_catalogo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            
            return response
            
        except ImportError:
            # Si WeasyPrint no está instalado, usar alternativa simple
            messages.warning(request, "Exportación PDF no disponible. Instale WeasyPrint para habilitar esta función.")
            return self._export_html(request)
    
    def _export_excel(self, request):
        """Exportar reporte a Excel"""
        from django.http import HttpResponse
        import io
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.chart import BarChart, Reference
            
            # Crear workbook
            wb = openpyxl.Workbook()
            
            # Obtener datos
            data = self._get_report_data()
            
            # Hoja 1: Resumen General
            ws_summary = wb.active
            ws_summary.title = "Resumen General"
            
            # Título
            ws_summary['A1'] = 'REPORTE DE CATÁLOGO - FORGEDB'
            ws_summary['A1'].font = Font(size=16, bold=True)
            ws_summary['A2'] = f'Generado: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
            
            # Estadísticas principales
            row = 4
            stats = [
                ('Total Items en Catálogo', data.get('total_catalog_items', 0)),
                ('Tipos de Equipo', data.get('equipment_types_count', 0)),
                ('Proveedores', data.get('suppliers_count', 0)),
                ('Proveedores Activos', data.get('active_suppliers_count', 0)),
                ('Códigos de Referencia', data.get('total_reference_codes', 0)),
                ('Sistemas de Taxonomía', data.get('taxonomy_systems_count', 0)),
                ('Subsistemas de Taxonomía', data.get('taxonomy_subsystems_count', 0)),
                ('Grupos de Taxonomía', data.get('taxonomy_groups_count', 0)),
                ('Monedas Configuradas', data.get('currencies_count', 0)),
            ]
            
            ws_summary['A' + str(row)] = 'Métrica'
            ws_summary['B' + str(row)] = 'Valor'
            ws_summary['A' + str(row)].font = Font(bold=True)
            ws_summary['B' + str(row)].font = Font(bold=True)
            
            for metric, value in stats:
                row += 1
                ws_summary['A' + str(row)] = metric
                ws_summary['B' + str(row)] = value
            
            # Ajustar anchos de columna
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 15
            
            # Hoja 2: Códigos de Referencia
            ws_codes = wb.create_sheet("Códigos de Referencia")
            ws_codes['A1'] = 'CÓDIGOS DE REFERENCIA POR CATEGORÍA'
            ws_codes['A1'].font = Font(size=14, bold=True)
            
            row = 3
            ws_codes['A' + str(row)] = 'Categoría'
            ws_codes['B' + str(row)] = 'Cantidad'
            ws_codes['A' + str(row)].font = Font(bold=True)
            ws_codes['B' + str(row)].font = Font(bold=True)
            
            reference_codes = data.get('reference_code_counts', {})
            for category, count in reference_codes.items():
                row += 1
                ws_codes['A' + str(row)] = category.title()
                ws_codes['B' + str(row)] = count
            
            ws_codes.column_dimensions['A'].width = 20
            ws_codes.column_dimensions['B'].width = 15
            
            # Hoja 3: Taxonomía
            ws_taxonomy = wb.create_sheet("Taxonomía")
            ws_taxonomy['A1'] = 'ESTRUCTURA DE TAXONOMÍA'
            ws_taxonomy['A1'].font = Font(size=14, bold=True)
            
            ws_taxonomy['A3'] = 'Nivel'
            ws_taxonomy['B3'] = 'Cantidad'
            ws_taxonomy['A3'].font = Font(bold=True)
            ws_taxonomy['B3'].font = Font(bold=True)
            
            ws_taxonomy['A4'] = 'Sistemas'
            ws_taxonomy['B4'] = data.get('taxonomy_systems_count', 0)
            ws_taxonomy['A5'] = 'Subsistemas'
            ws_taxonomy['B5'] = data.get('taxonomy_subsystems_count', 0)
            ws_taxonomy['A6'] = 'Grupos'
            ws_taxonomy['B6'] = data.get('taxonomy_groups_count', 0)
            
            ws_taxonomy.column_dimensions['A'].width = 20
            ws_taxonomy.column_dimensions['B'].width = 15
            
            # Guardar en memoria
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Crear respuesta HTTP
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="reporte_catalogo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            return response
            
        except ImportError:
            messages.error(request, "Exportación Excel no disponible. Instale openpyxl para habilitar esta función.")
            return redirect('frontend:catalog_reports')
    
    def _export_html(self, request):
        """Exportar reporte como HTML simple (fallback)"""
        from django.http import HttpResponse
        from django.template.loader import render_to_string
        
        context = self._get_report_data()
        context['export_date'] = timezone.now()
        
        html_string = render_to_string('frontend/catalog/catalog_report_simple.html', context)
        
        response = HttpResponse(html_string, content_type='text/html')
        response['Content-Disposition'] = f'attachment; filename="reporte_catalogo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.html"'
        
        return response
    
    def _get_report_data(self):
        """Obtener datos del reporte"""
        from django.utils import timezone
        
        context = {}
        
        try:
            api_client = self.get_api_client()
            
            # Estadísticas de Tipos de Equipo
            try:
                equipment_types_data = api_client.get('equipment-types/', params={'page_size': 1})
                context['equipment_types_count'] = equipment_types_data.get('count', 0)
            except APIException:
                context['equipment_types_count'] = 0
            
            # Estadísticas de Taxonomía
            try:
                taxonomy_systems = api_client.get('taxonomy-systems/', params={'page_size': 1})
                context['taxonomy_systems_count'] = taxonomy_systems.get('count', 0)
                
                taxonomy_subsystems = api_client.get('taxonomy-subsystems/', params={'page_size': 1})
                context['taxonomy_subsystems_count'] = taxonomy_subsystems.get('count', 0)
                
                taxonomy_groups = api_client.get('taxonomy-groups/', params={'page_size': 1})
                context['taxonomy_groups_count'] = taxonomy_groups.get('count', 0)
            except APIException:
                context['taxonomy_systems_count'] = 0
                context['taxonomy_subsystems_count'] = 0
                context['taxonomy_groups_count'] = 0
            
            # Estadísticas de Códigos de Referencia
            reference_code_counts = {}
            for category in ['fuel', 'transmission', 'color', 'drivetrain', 'condition', 'aspiration']:
                try:
                    codes_data = api_client.get(f'{category}-codes/', params={'page_size': 1})
                    reference_code_counts[category] = codes_data.get('count', 0)
                except APIException:
                    reference_code_counts[category] = 0
            
            context['reference_code_counts'] = reference_code_counts
            context['total_reference_codes'] = sum(reference_code_counts.values())
            
            # Estadísticas de Monedas
            try:
                currencies_data = api_client.get('currencies/', params={'page_size': 1})
                context['currencies_count'] = currencies_data.get('count', 0)
            except APIException:
                context['currencies_count'] = 0
            
            # Estadísticas de Proveedores
            try:
                suppliers_data = api_client.get('suppliers/', params={'page_size': 1})
                context['suppliers_count'] = suppliers_data.get('count', 0)
                
                active_suppliers = api_client.get('suppliers/', params={
                    'page_size': 1,
                    'is_active': 'true'
                })
                context['active_suppliers_count'] = active_suppliers.get('count', 0)
            except APIException:
                context['suppliers_count'] = 0
                context['active_suppliers_count'] = 0
            
            # Resumen general
            context['total_catalog_items'] = (
                context['equipment_types_count'] +
                context['taxonomy_systems_count'] +
                context['taxonomy_subsystems_count'] +
                context['taxonomy_groups_count'] +
                context['total_reference_codes'] +
                context['currencies_count'] +
                context['suppliers_count']
            )
            
        except Exception as e:
            logger.error(f"Error al obtener datos del reporte: {str(e)}")
        
        return context


# =============================================================================
# Category Management Views
# =============================================================================

class CategoryListView(LoginRequiredMixin, APIClientMixin, TemplateView):
    """Category list view"""
    template_name = 'frontend/catalog/category_list.html'
    login_url = 'frontend:login'
    paginate_by = 20
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get search and filter parameters
        search = self.request.GET.get('search', '').strip()
        page = self._get_page_number()
        status_filter = self.request.GET.get('status', '')
        sort_by = self.request.GET.get('sort', 'sort_order')
        sort_order = self.request.GET.get('order', 'asc')
        
        try:
            api_client = self.get_api_client()
            
            # Build filter parameters
            filters = {'page': page, 'page_size': self.paginate_by}
            if search:
                filters['search'] = search
            if status_filter:
                filters['is_active'] = status_filter == 'active'
            order_prefix = '-' if sort_order == 'desc' else ''
            filters['ordering'] = f"{order_prefix}{sort_by}"
            
            # Get categories data
            categories_data = api_client.get('categories/', params=filters)
            context['categories'] = categories_data.get('results', [])
            
            # Process categories for display
            for category in context['categories']:
                if category.get('is_active'):
                    category['status_class'] = 'success'
                    category['status_label'] = 'Activa'
                else:
                    category['status_class'] = 'secondary'
                    category['status_label'] = 'Inactiva'
                
                # Equipment type count
                category['equipment_count'] = category.get('equipment_type_count', 0)
            
            # Pagination
            total_count = categories_data.get('count', 0)
            total_pages = (total_count + self.paginate_by - 1) // self.paginate_by
            
            context['pagination'] = {
                'count': total_count,
                'next': categories_data.get('next'),
                'previous': categories_data.get('previous'),
                'current_page': page,
                'total_pages': total_pages,
                'page_range': self._get_page_range(page, total_pages),
                'has_previous': page > 1,
                'has_next': page < total_pages,
                'start_index': (page - 1) * self.paginate_by + 1,
                'end_index': min(page * self.paginate_by, total_count),
            }
            
            # Filter context
            context['filters'] = {
                'search': search,
                'status': status_filter,
                'sort': sort_by,
                'order': sort_order,
            }
            
        except APIException as e:
            self.handle_api_error(e, "Error al cargar categorías")
            context['categories'] = []
            context['pagination'] = self._get_empty_pagination()
            context['filters'] = {
                'search': search,
                'status': status_filter,
                'sort': sort_by,
                'order': sort_order,
            }
        
        return context
    
    def _get_page_number(self):
        try:
            page = int(self.request.GET.get('page', 1))
            return max(1, page)
        except (ValueError, TypeError):
            return 1
    
    def _get_page_range(self, current_page, total_pages, window=5):
        if total_pages <= window:
            return list(range(1, total_pages + 1))
        half_window = window // 2
        start = max(1, current_page - half_window)
        end = min(total_pages, current_page + half_window)
        if end - start < window - 1:
            if start == 1:
                end = min(total_pages, start + window - 1)
            else:
                start = max(1, end - window + 1)
        return list(range(start, end + 1))
    
    def _get_empty_pagination(self):
        return {
            'count': 0,
            'current_page': 1,
            'total_pages': 0,
            'page_range': [],
            'has_previous': False,
            'has_next': False,
            'start_index': 0,
            'end_index': 0,
        }


class CategoryDetailView(LoginRequiredMixin, APIClientMixin, TemplateView):
    """Category detail view"""
    template_name = 'frontend/catalog/category_detail.html'
    login_url = 'frontend:login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        category_id = kwargs.get('pk')
        
        try:
            api_client = self.get_api_client()
            category = api_client.get(f'categories/{category_id}/')
            context['category'] = category
            
            # Process status
            if category.get('is_active'):
                category['status_class'] = 'success'
                category['status_label'] = 'Activa'
            else:
                category['status_class'] = 'secondary'
                category['status_label'] = 'Inactiva'
            
            # Get equipment types in this category
            try:
                equipment_types = api_client.get('equipment-types/', params={
                    'category': category_id,
                    'page_size': 50
                })
                context['equipment_types'] = equipment_types.get('results', [])
            except APIException:
                context['equipment_types'] = []
            
        except APIException as e:
            self.handle_api_error(e, "Error al cargar categoría")
            context['category'] = None
        
        return context


class CategoryCreateView(LoginRequiredMixin, APIClientMixin, TemplateView):
    """Category create view"""
    template_name = 'frontend/catalog/category_form.html'
    login_url = 'frontend:login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear Categoría'
        context['action'] = 'create'
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            # Build category data
            category_data = {
                'category_code': request.POST.get('category_code', '').strip().upper(),
                'name': request.POST.get('name', '').strip(),
                'description': request.POST.get('description', '').strip(),
                'icon': request.POST.get('icon', '').strip(),
                'color': request.POST.get('color', '').strip(),
                'sort_order': int(request.POST.get('sort_order', 0)),
                'is_active': request.POST.get('is_active') == 'on',
            }
            
            # Validate required fields
            if not category_data['category_code']:
                messages.error(request, "El código de categoría es obligatorio")
                return self._render_form_with_errors(category_data)
            if not category_data['name']:
                messages.error(request, "El nombre de categoría es obligatorio")
                return self._render_form_with_errors(category_data)
            
            # Try API client first
            try:
                api_client = self.get_api_client()
                result = api_client.post('categories/', category_data)
                messages.success(request, f"Categoría '{result['name']}' creada exitosamente")
                return redirect('frontend:category_list')
            except APIException as e:
                # If API fails due to auth, try direct Django ORM
                if e.status_code == 401:
                    logger.info("API authentication failed, falling back to direct ORM")
                    from core.models import Category
                    category = Category.objects.create(**category_data)
                    messages.success(request, f"Categoría '{category.name}' creada exitosamente")
                    return redirect('frontend:category_list')
                else:
                    # Re-raise other API exceptions
                    raise e
            
        except APIException as e:
            return self._render_form_with_errors(request.POST, str(e))
        except Exception as e:
            logger.error(f"Unexpected error in category creation: {e}", exc_info=True)
            return self._render_form_with_errors(request.POST, "Error inesperado al crear la categoría")
    
    def _render_form_with_errors(self, data, error=None):
        context = self.get_context_data()
        context['form_data'] = data
        if error:
            messages.error(self.request, error)
        return self.render_to_response(context)


class CategoryUpdateView(LoginRequiredMixin, APIClientMixin, TemplateView):
    """Category update view"""
    template_name = 'frontend/catalog/category_form.html'
    login_url = 'frontend:login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        category_id = kwargs.get('pk')
        
        try:
            api_client = self.get_api_client()
            category = api_client.get(f'categories/{category_id}/')
            context['title'] = 'Editar Categoría'
            context['action'] = 'update'
            context['category'] = category
            context['form_data'] = {
                'category_code': category.get('category_code', ''),
                'name': category.get('name', ''),
                'description': category.get('description', ''),
                'icon': category.get('icon', ''),
                'color': category.get('color', ''),
                'sort_order': category.get('sort_order', 0),
                'is_active': category.get('is_active', True),
            }
        except APIException as e:
            self.handle_api_error(e, "Error al cargar categoría")
            return redirect('frontend:category_list')
        
        return context
    
    def post(self, request, *args, **kwargs):
        category_id = kwargs.get('pk')
        
        try:
            # Build category data
            category_data = {
                'category_code': request.POST.get('category_code', '').strip().upper(),
                'name': request.POST.get('name', '').strip(),
                'description': request.POST.get('description', '').strip(),
                'icon': request.POST.get('icon', '').strip(),
                'color': request.POST.get('color', '').strip(),
                'sort_order': int(request.POST.get('sort_order', 0)),
                'is_active': request.POST.get('is_active') == 'on',
            }
            
            # Validate required fields
            if not category_data['category_code']:
                messages.error(request, "El código de categoría es obligatorio")
                return self._render_form_with_errors(category_data)
            if not category_data['name']:
                messages.error(request, "El nombre de categoría es obligatorio")
                return self._render_form_with_errors(category_data)
            
            # Try API client first
            try:
                api_client = self.get_api_client()
                result = api_client.put(f'categories/{category_id}/', category_data)
                messages.success(request, f"Categoría '{result['name']}' actualizada exitosamente")
                return redirect('frontend:category_list')
            except APIException as e:
                # If API fails due to auth, try direct Django ORM
                if e.status_code == 401:
                    logger.info("API authentication failed, falling back to direct ORM for update")
                    from core.models import Category
                    category = Category.objects.get(pk=category_id)
                    for key, value in category_data.items():
                        setattr(category, key, value)
                    category.save()
                    messages.success(request, f"Categoría '{category.name}' actualizada exitosamente")
                    return redirect('frontend:category_list')
                else:
                    # Re-raise other API exceptions
                    raise e
            
        except APIException as e:
            return self._render_form_with_errors(request.POST, str(e))
        except Exception as e:
            logger.error(f"Unexpected error in category update: {e}", exc_info=True)
            return self._render_form_with_errors(request.POST, "Error inesperado al actualizar la categoría")
    
    def _render_form_with_errors(self, data, error=None):
        context = self.get_context_data()
        context['form_data'] = data
        if error:
            messages.error(self.request, error)
        return self.render_to_response(context)


class CategoryToggleActiveView(LoginRequiredMixin, APIClientMixin, View):
    """Toggle category active status (soft delete)"""
    login_url = 'frontend:login'
    
    def post(self, request, *args, **kwargs):
        category_id = kwargs.get('pk')
        
        try:
            # Try API client first
            try:
                api_client = self.get_api_client()
                
                # Get current category status
                category = api_client.get(f'categories/{category_id}/')
                current_status = category.get('is_active', True)
                new_status = not current_status
                
                # Update status
                data = {'is_active': new_status}
                result = api_client.patch(f'categories/{category_id}/', data)
                
                status_text = "activada" if new_status else "desactivada"
                messages.success(request, f"Categoría '{category['name']}' {status_text} exitosamente")
                
            except APIException as e:
                # If API fails due to auth, try direct Django ORM
                if e.status_code == 401:
                    logger.info("API authentication failed, falling back to direct ORM for toggle")
                    from core.models import Category
                    category = Category.objects.get(pk=category_id)
                    
                    # Toggle status
                    category.is_active = not category.is_active
                    category.save()
                    
                    status_text = "activada" if category.is_active else "desactivada"
                    messages.success(request, f"Categoría '{category.name}' {status_text} exitosamente")
                else:
                    self.handle_api_error(e, "Error al cambiar estado de categoría")
        
        except Exception as e:
            logger.error(f"Error toggling category status: {e}")
            messages.error(request, "Error al cambiar estado de categoría")
        
        return redirect('frontend:category_list')


class CategoryDeleteView(LoginRequiredMixin, APIClientMixin, View):
    """Category delete view - DBA ONLY via direct URL access"""
    login_url = 'frontend:login'
    
    def dispatch(self, request, *args, **kwargs):
        # Check if user has DBA role
        if not hasattr(request.user, 'technicianuser') or not request.user.technicianuser.is_dba:
            messages.error(request, "Solo usuarios DBA pueden eliminar categorías permanentemente")
            return redirect('frontend:category_list')
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request, *args, **kwargs):
        """Show confirmation page for DBA users"""
        category_id = kwargs.get('pk')
        try:
            api_client = self.get_api_client()
            category = api_client.get(f'categories/{category_id}/')
            context = {
                'category': category,
                'title': f'Eliminar Categoría: {category["name"]}'
            }
            return render(request, 'frontend/catalog/category_delete_confirm.html', context)
        except APIException as e:
            self.handle_api_error(e, "Error al cargar categoría para eliminación")
            return redirect('frontend:category_list')
    
    def post(self, request, *args, **kwargs):
        category_id = kwargs.get('pk')
        
        try:
            # Try API client first
            try:
                api_client = self.get_api_client()
                
                # Check if category has equipment types
                category = api_client.get(f'categories/{category_id}/')
                equipment_count = category.get('equipment_type_count', 0)
                
                if equipment_count > 0:
                    messages.warning(
                        request,
                        f"No se puede eliminar la categoría '{category['name']}' porque tiene {equipment_count} tipos de equipo asociados"
                    )
                    return redirect('frontend:category_list')
                
                # Delete category
                api_client.delete(f'categories/{category_id}/')
                messages.success(request, f"Categoría '{category['name']}' eliminada exitosamente")
                
            except APIException as e:
                # If API fails due to auth, try direct Django ORM
                if e.status_code == 401:
                    logger.info("API authentication failed, falling back to direct ORM for delete")
                    from core.models import Category
                    category = Category.objects.get(pk=category_id)
                    
                    # Check if category has equipment types
                    equipment_count = category.equipmenttype_set.count()
                    if equipment_count > 0:
                        messages.warning(
                            request,
                            f"No se puede eliminar la categoría '{category.name}' porque tiene {equipment_count} tipos de equipo asociados"
                        )
                        return redirect('frontend:category_list')
                    
                    # Delete category
                    category_name = category.name
                    category.delete()
                    messages.success(request, f"Categoría '{category_name}' eliminada exitosamente")
                else:
                    error_msg = str(e)
                    # Check if it's a foreign key constraint error
                    if 'foreign key' in error_msg.lower() or 'associated' in error_msg.lower():
                        messages.warning(request, "No se puede eliminar esta categoría porque tiene datos asociados")
                    else:
                        self.handle_api_error(e, "Error al eliminar categoría")
        
        except Exception as e:
            logger.error(f"Error deleting category: {e}")
            messages.error(request, "Error al eliminar categoría")
        
        return redirect('frontend:category_list')
