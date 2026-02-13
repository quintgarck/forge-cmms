"""
Equipment Type Views - Frontend CRUD Operations
ForgeDB Frontend Web Application

This module implements complete CRUD operations for Equipment Types
with Django Class-Based Views, form validation, and API integration.
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DetailView, DeleteView, TemplateView, View
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, Http404, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django import forms
from django.utils import timezone
import json
import logging

from ..services.api_client import ForgeAPIClient, APIException
from ..forms.equipment_type_forms import EquipmentTypeForm, EquipmentTypeSearchForm
from ..mixins import APIClientMixin

logger = logging.getLogger(__name__)


class EquipmentTypeListView(LoginRequiredMixin, APIClientMixin, ListView):
    """
    Lista paginada de tipos de equipo con búsqueda y filtros
    """
    template_name = 'frontend/catalog/equipment_type_list.html'
    context_object_name = 'equipment_types'
    paginate_by = 20
    
    def dispatch(self, request, *args, **kwargs):
        # Clear cache for equipment types on every list view access
        # This ensures newly added items appear immediately
        try:
            from django.core.cache import cache
            # Limpiar todas las posibles claves de caché
            cache.delete('forge_api:equipment-types')
            cache.delete('forge_api:equipment-types_')
            cache.delete('forge_api:equipment-types__page_1_page_size_20')
            # Intentar limpiar con prefijo si el backend lo soporta
            try:
                keys = cache._cache.keys() if hasattr(cache, '_cache') else []
                for key in list(keys):
                    if 'equipment-types' in str(key):
                        cache.delete(key)
            except:
                pass
        except Exception:
            pass
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        """Obtener datos de tipos de equipo desde la API"""
        try:
            api_client = self.get_api_client()
            
            # Parámetros de búsqueda y filtrado
            params = {}
            
            # Búsqueda por texto
            search = self.request.GET.get('search', '').strip()
            if search:
                params['search'] = search
            
            # Filtro por categoría
            category = self.request.GET.get('category', '').strip()
            if category:
                params['category'] = category
            
            # Filtro por estado activo
            is_active = self.request.GET.get('is_active', '').strip()
            if is_active:
                params['is_active'] = is_active
            
            # Paginación
            page = self.request.GET.get('page', 1)
            params['page'] = page
            params['page_size'] = self.paginate_by
            
            # Llamada a la API usando el método específico
            response = api_client.get_equipment_types(**params)
            
            if response and 'results' in response:
                return response['results']
            else:
                messages.error(self.request, "Error al cargar los tipos de equipo.")
                return []
                
        except APIException as e:
            logger.error(f"API error loading equipment types: {str(e)}")
            self.handle_api_error(e, "Error al cargar los tipos de equipo.")
            return []
        except Exception as e:
            logger.error(f"Unexpected error loading equipment types: {str(e)}")
            messages.error(self.request, "Error de conexión con el servidor.")
            return []
    
    def get_context_data(self, **kwargs):
        """Agregar contexto adicional para la vista"""
        context = super().get_context_data(**kwargs)
        
        # Obtener categorías dinámicas
        categories_choices = []
        try:
            api_client = self.get_api_client()
            categories_response = api_client.get('categories/', params={'is_active': True, 'ordering': 'sort_order,name'})
            if categories_response and 'results' in categories_response:
                categories_choices = [(cat['category_code'], cat['name']) for cat in categories_response['results']]
        except Exception as e:
            logger.warning(f"Error loading categories for search form: {str(e)}. Using fallback.")
            # Fallback si hay error
            categories_choices = [
                ('AUTOMOTRIZ', 'Automotriz'),
                ('INDUSTRIAL', 'Industrial'),
                ('AGRICOLA', 'Agrícola'),
                ('CONSTRUCCION', 'Construcción'),
                ('ELECTRONICO', 'Electrónico'),
                ('CAT001', 'CAT TRX'),
                ('OTRO', 'Otro'),
            ]
        
        # Formulario de búsqueda
        context['search_form'] = EquipmentTypeSearchForm(self.request.GET, categories_choices=categories_choices)
        
        # Parámetros actuales para mantener en paginación
        context['current_search'] = self.request.GET.get('search', '')
        context['current_category'] = self.request.GET.get('category', '')
        context['current_is_active'] = self.request.GET.get('is_active', '')
        
        # Categorías disponibles para el filtro - obtenidas desde la base de datos
        try:
            api_client = self.get_api_client()
            categories_response = api_client.get('categories/', params={'is_active': True, 'ordering': 'sort_order,name'})
            if categories_response and 'results' in categories_response:
                context['categories'] = [
                    (cat['category_code'], cat['name']) 
                    for cat in categories_response['results']
                ]
                # Convertir al formato esperado por la plantilla
                context['category_options'] = [
                    {'value': '', 'label': 'Todas las categorías'},
                ] + [
                    {'value': cat['category_code'], 'label': cat['name']} 
                    for cat in categories_response['results']
                ]
            else:
                # Fallback si la API falla - usando categorías reales de la BD
                context['categories'] = [
                    ('AUTOMOTRIZ', 'Automotriz'),
                    ('INDUSTRIAL', 'Industrial'),
                    ('AGRICOLA', 'Agrícola'),
                    ('CONSTRUCCION', 'Construcción'),
                    ('ELECTRONICO', 'Electrónico'),
                    ('CAT001', 'CAT TRX'),
                    ('OTRO', 'Otro'),
                ]
                # Convertir al formato esperado por la plantilla
                context['category_options'] = [
                    {'value': '', 'label': 'Todas las categorías'},
                ] + [
                    {'value': code, 'label': name} 
                    for code, name in context['categories']
                ]
        except Exception as e:
            logger.warning(f"Error loading categories from API: {str(e)}. Using fallback.")
            # Fallback si hay error de conexión - usando categorías reales de la BD
            context['categories'] = [
                ('AUTOMOTRIZ', 'Automotriz'),
                ('INDUSTRIAL', 'Industrial'),
                ('AGRICOLA', 'Agrícola'),
                ('CONSTRUCCION', 'Construcción'),
                ('ELECTRONICO', 'Electrónico'),
                ('CAT001', 'CAT TRX'),
                ('OTRO', 'Otro'),
            ]
            # Convertir al formato esperado por la plantilla
            context['category_options'] = [
                {'value': '', 'label': 'Todas las categorías'},
            ] + [
                {'value': code, 'label': name} 
                for code, name in context['categories']
            ]
        
        # Información de paginación
        try:
            api_client = self.get_api_client()
            params = {
                'search': self.request.GET.get('search', ''),
                'category': self.request.GET.get('category', ''),
                'is_active': self.request.GET.get('is_active', ''),
                'page': self.request.GET.get('page', 1),
                'page_size': self.paginate_by
            }
            response = api_client.get_equipment_types(**params)
            
            if response:
                context['total_count'] = response.get('count', 0)
                context['has_next'] = response.get('next') is not None
                context['has_previous'] = response.get('previous') is not None
                context['current_page'] = int(self.request.GET.get('page', 1))
                
                # Calcular páginas para navegación
                total_pages = (context['total_count'] + self.paginate_by - 1) // self.paginate_by
                context['total_pages'] = total_pages
                
        except Exception as e:
            logger.error(f"Error getting pagination info: {str(e)}")
            context['total_count'] = 0
            context['has_next'] = False
            context['has_previous'] = False
            context['current_page'] = 1
            context['total_pages'] = 1
        
        return context


class EquipmentTypeCreateView(LoginRequiredMixin, APIClientMixin, TemplateView):
    """
    Vista para crear nuevos tipos de equipo
    """
    template_name = 'frontend/catalog/equipment_type_form.html'
    login_url = 'frontend:login'
    
    def _get_categories(self):
        """Obtener las categorías completas desde la API (code -> id mapping)"""
        try:
            api_client = self.get_api_client()
            categories_response = api_client.get('categories/', params={'is_active': True, 'ordering': 'sort_order,name'})
            if categories_response and 'results' in categories_response:
                # Return list of categories with code, name, and id
                return categories_response['results']
        except Exception as e:
            logger.warning(f"Error loading categories: {str(e)}. Using fallback.")
        
        # Fallback si hay error
        return [
            {'category_code': 'AUTOMOTRIZ', 'name': 'Automotriz', 'category_id': 1},
            {'category_code': 'INDUSTRIAL', 'name': 'Industrial', 'category_id': 2},
            {'category_code': 'AGRICOLA', 'name': 'Agrícola', 'category_id': 3},
        ]
    
    def _get_categories_choices(self):
        """Obtener las categorías como choices para el formulario"""
        categories = self._get_categories()
        return [(cat['category_code'], cat['name']) for cat in categories]
    
    def _get_category_id_from_code(self, code):
        """Convertir código de categoría a ID numérico"""
        categories = self._get_categories()
        for cat in categories:
            if cat['category_code'] == code:
                return cat.get('category_id') or cat.get('id')
        return None
    
    def get_context_data(self, **kwargs):
        """Agregar contexto para la vista de creación"""
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear Tipo de Equipo'
        context['submit_text'] = 'Crear Tipo de Equipo'
        context['cancel_url'] = reverse('frontend:equipment_type_list')
        
        # Obtener categorías dinámicas
        categories_choices = self._get_categories_choices()
        
        # Agregar formulario al contexto
        if self.request.method == 'POST':
            context['form'] = EquipmentTypeForm(self.request.POST, categories_choices=categories_choices)
        else:
            context['form'] = EquipmentTypeForm(categories_choices=categories_choices)
        
        return context
    
    def post(self, request, *args, **kwargs):
        """Manejar solicitud POST para crear tipo de equipo"""
        # Obtener categorías dinámicas para el formulario
        categories_choices = self._get_categories_choices()
        form = EquipmentTypeForm(request.POST, categories_choices=categories_choices)
        
        if form.is_valid():
            try:
                api_client = self.get_api_client()
                
                # Convertir código de categoría a ID
                category_code = form.cleaned_data['category']
                category_id = self._get_category_id_from_code(category_code)
                
                if not category_id:
                    messages.error(request, f"Categoría '{category_code}' no válida.")
                    context = self.get_context_data(**kwargs)
                    context['form'] = form
                    return self.render_to_response(context)
                
                # Preparar datos para la API
                data = {
                    'type_code': form.cleaned_data['type_code'].upper(),
                    'category': category_id,
                    'name': form.cleaned_data['name'],
                    'icon': form.cleaned_data.get('icon', ''),
                    'color': form.cleaned_data.get('color', ''),
                    'attr_schema': form.cleaned_data.get('attr_schema', {}),
                    'description': form.cleaned_data.get('description', ''),
                    'is_active': form.cleaned_data.get('is_active', True)
                }
                
                # Enviar a la API usando el método específico
                response = api_client.create_equipment_type(data)
                
                if response:
                    # Limpiar caché agresivamente para asegurar que aparezca en la lista
                    from django.core.cache import cache
                    cache.delete('forge_api:equipment-types')
                    cache.delete('forge_api:equipment-types_')
                    
                    messages.success(
                        request,
                        f"Tipo de equipo '{data['name']}' creado exitosamente (ID: {response.get('type_id', 'N/A')})."
                    )
                    return redirect('frontend:equipment_type_list')
                else:
                    messages.error(
                        request, 
                        "Error al crear el tipo de equipo. Verifica los datos ingresados."
                    )
                    
            except APIException as e:
                logger.error(f"API error creating equipment type: {str(e)}")
                self.handle_api_error(e, "Error al crear el tipo de equipo.")
            except Exception as e:
                logger.error(f"Unexpected error creating equipment type: {str(e)}")
                messages.error(
                    request, 
                    "Error de conexión con el servidor. Intenta nuevamente."
                )
        else:
            messages.error(
                request, 
                "Por favor corrige los errores en el formulario."
            )
        
        # Si hay errores, volver a mostrar el formulario con los datos
        context = self.get_context_data(**kwargs)
        context['form'] = form
        return self.render_to_response(context)
    
    def form_valid(self, form):
        """Procesar formulario válido enviando datos a la API"""
        try:
            api_client = self.get_api_client()
            
            # Preparar datos para la API
            data = {
                'type_code': form.cleaned_data['type_code'].upper(),
                'category': form.cleaned_data['category'],
                'name': form.cleaned_data['name'],
                'icon': form.cleaned_data.get('icon', ''),
                'color': form.cleaned_data.get('color', ''),
                'attr_schema': form.cleaned_data.get('attr_schema', {}),
                'description': form.cleaned_data.get('description', ''),
                'is_active': form.cleaned_data.get('is_active', True)
            }
            
            # Enviar a la API usando el método específico
            response = api_client.create_equipment_type(data)
            
            if response:
                messages.success(
                    self.request, 
                    f"Tipo de equipo '{data['name']}' creado exitosamente."
                )
                return redirect(self.success_url)
            else:
                messages.error(
                    self.request, 
                    "Error al crear el tipo de equipo. Verifica los datos ingresados."
                )
                return self.form_invalid(form)
                
        except APIException as e:
            logger.error(f"API error creating equipment type: {str(e)}")
            self.handle_api_error(e, "Error al crear el tipo de equipo.")
            return self.form_invalid(form)
        except Exception as e:
            logger.error(f"Unexpected error creating equipment type: {str(e)}")
            messages.error(
                self.request, 
                "Error de conexión con el servidor. Intenta nuevamente."
            )
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        """Manejar formulario inválido"""
        messages.error(
            self.request, 
            "Por favor corrige los errores en el formulario."
        )
        return super().form_invalid(form)


class EquipmentTypeToggleActiveView(LoginRequiredMixin, APIClientMixin, View):
    """
    Vista para alternar el estado activo/inactivo de tipos de equipo
    """
    login_url = 'frontend:login'
    
    def post(self, request, pk):
        """Alternar estado activo del tipo de equipo"""
        try:
            api_client = self.get_api_client()
            
            # Obtener el tipo de equipo actual
            equipment_type = api_client.get_equipment_type(pk)
            
            if not equipment_type:
                messages.error(request, "Tipo de equipo no encontrado.")
                return redirect('frontend:equipment_type_list')
            
            # Alternar el estado
            current_status = equipment_type.get('is_active', True)
            new_status = not current_status
            
            # Actualizar el tipo de equipo
            data = {
                'is_active': new_status
            }
            
            response = api_client.update_equipment_type(pk, data)
            
            if response:
                status_text = "activado" if new_status else "desactivado"
                messages.success(
                    request, 
                    f"Tipo de equipo '{equipment_type.get('name')}' {status_text} exitosamente."
                )
            else:
                messages.error(
                    request, 
                    "Error al actualizar el estado del tipo de equipo."
                )
                
        except APIException as e:
            logger.error(f"API error toggling equipment type {pk}: {str(e)}")
            self.handle_api_error(e, "Error al cambiar el estado del tipo de equipo.")
        except Exception as e:
            logger.error(f"Unexpected error toggling equipment type {pk}: {str(e)}")
            messages.error(
                request, 
                "Error al cambiar el estado del tipo de equipo. Intenta nuevamente."
            )
        
        return redirect('frontend:equipment_type_list')


class EquipmentTypeUpdateView(LoginRequiredMixin, APIClientMixin, TemplateView):
    """
    Vista para editar tipos de equipo existentes
    """
    template_name = 'frontend/catalog/equipment_type_form.html'
    login_url = 'frontend:login'
    
    def _get_categories(self):
        """Obtener las categorías completas desde la API (code -> id mapping)"""
        try:
            api_client = self.get_api_client()
            categories_response = api_client.get('categories/', params={'is_active': True, 'ordering': 'sort_order,name'})
            if categories_response and 'results' in categories_response:
                return categories_response['results']
        except Exception as e:
            logger.warning(f"Error loading categories: {str(e)}. Using fallback.")
        
        # Fallback si hay error
        return [
            {'category_code': 'AUTOMOTRIZ', 'name': 'Automotriz', 'category_id': 1},
            {'category_code': 'INDUSTRIAL', 'name': 'Industrial', 'category_id': 2},
            {'category_code': 'AGRICOLA', 'name': 'Agrícola', 'category_id': 3},
        ]
    
    def _get_categories_choices(self):
        """Obtener las categorías como choices para el formulario"""
        categories = self._get_categories()
        return [(cat['category_code'], cat['name']) for cat in categories]
    
    def _get_category_id_from_code(self, code):
        """Convertir código de categoría a ID numérico"""
        categories = self._get_categories()
        for cat in categories:
            if cat['category_code'] == code:
                return cat.get('category_id') or cat.get('id')
        return None
    
    def get_equipment_type(self):
        """Obtener el objeto desde la API"""
        try:
            api_client = self.get_api_client()
            type_id = self.kwargs.get('pk')
            
            response = api_client.get_equipment_type(type_id)
            
            if response:
                return response
            else:
                raise Http404("Tipo de equipo no encontrado")
                
        except APIException as e:
            logger.error(f"API error loading equipment type {self.kwargs.get('pk')}: {str(e)}")
            if e.status_code == 404:
                raise Http404("Tipo de equipo no encontrado")
            raise Http404("Error al cargar el tipo de equipo")
        except Exception as e:
            logger.error(f"Unexpected error loading equipment type {self.kwargs.get('pk')}: {str(e)}")
            raise Http404("Error al cargar el tipo de equipo")
    
    def get_context_data(self, **kwargs):
        """Agregar contexto para la vista de edición"""
        context = super().get_context_data(**kwargs)
        obj = self.get_equipment_type()
        context['title'] = f'Editar Tipo de Equipo: {obj.get("name", "")}'
        context['submit_text'] = 'Actualizar Tipo de Equipo'
        context['cancel_url'] = reverse('frontend:equipment_type_list')
        context['object'] = obj
        
        # Agregar formulario al contexto
        categories_choices = self._get_categories_choices()
        
        if self.request.method == 'POST':
            context['form'] = EquipmentTypeForm(self.request.POST, categories_choices=categories_choices)
        else:
            # Pre-poblar formulario con datos existentes
            import json
            
            # Convert attr_schema dict to JSON string if it's a dict
            attr_schema = obj.get('attr_schema', {})
            if isinstance(attr_schema, dict):
                attr_schema_json = json.dumps(attr_schema, ensure_ascii=False)
            else:
                attr_schema_json = attr_schema or '{}'
            
            # La API devuelve category como ID, necesitamos convertirlo a código
            category_value = obj.get('category', '')
            # Si es un número (ID), buscar el código correspondiente
            if isinstance(category_value, int) or (isinstance(category_value, str) and category_value.isdigit()):
                categories = self._get_categories()
                for cat in categories:
                    cat_id = str(cat.get('category_id') or cat.get('id', ''))
                    if cat_id == str(category_value):
                        category_value = cat['category_code']
                        break
            
            initial_data = {
                'type_code': obj.get('type_code', ''),
                'category': category_value,
                'name': obj.get('name', ''),
                'icon': obj.get('icon', ''),
                'color': obj.get('color', ''),
                'attr_schema': attr_schema_json,
                'description': obj.get('description', ''),
                'is_active': obj.get('is_active', True)
            }
            context['form'] = EquipmentTypeForm(initial=initial_data, categories_choices=categories_choices)
        
        return context
    
    def post(self, request, *args, **kwargs):
        """Manejar solicitud POST para actualizar tipo de equipo"""
        categories_choices = self._get_categories_choices()
        form = EquipmentTypeForm(request.POST, categories_choices=categories_choices)
        
        if form.is_valid():
            try:
                api_client = self.get_api_client()
                type_id = self.kwargs.get('pk')
                
                # Convertir código de categoría a ID
                category_code = form.cleaned_data['category']
                category_id = self._get_category_id_from_code(category_code)
                
                if not category_id:
                    messages.error(request, f"Categoría '{category_code}' no válida.")
                    context = self.get_context_data(**kwargs)
                    context['form'] = form
                    return self.render_to_response(context)
                
                # Preparar datos para la API
                data = {
                    'type_code': form.cleaned_data['type_code'].upper(),
                    'category': category_id,
                    'name': form.cleaned_data['name'],
                    'icon': form.cleaned_data.get('icon', ''),
                    'color': form.cleaned_data.get('color', ''),
                    'attr_schema': form.cleaned_data.get('attr_schema', {}),
                    'description': form.cleaned_data.get('description', ''),
                    'is_active': form.cleaned_data.get('is_active', True)
                }
                
                # Enviar actualización a la API
                response = api_client.update_equipment_type(type_id, data)
                
                if response:
                    # Limpiar caché para reflejar cambios en la lista
                    from django.core.cache import cache
                    cache.delete('forge_api:equipment-types')
                    cache.delete('forge_api:equipment-types_')
                    
                    messages.success(
                        request,
                        f"Tipo de equipo '{data['name']}' actualizado exitosamente."
                    )
                    return redirect('frontend:equipment_type_list')
                else:
                    messages.error(
                        request, 
                        "Error al actualizar el tipo de equipo. Verifica los datos ingresados."
                    )
                    
            except APIException as e:
                logger.error(f"API error updating equipment type {self.kwargs.get('pk')}: {str(e)}")
                self.handle_api_error(e, "Error al actualizar el tipo de equipo.")
            except Exception as e:
                logger.error(f"Unexpected error updating equipment type {self.kwargs.get('pk')}: {str(e)}")
                messages.error(
                    request, 
                    "Error de conexión con el servidor. Intenta nuevamente."
                )
        else:
            messages.error(
                request, 
                "Por favor corrige los errores en el formulario."
            )
        
        # Si hay errores, volver a mostrar el formulario con los datos
        context = self.get_context_data(**kwargs)
        context['form'] = form
        return self.render_to_response(context)


class EquipmentTypeDetailView(LoginRequiredMixin, APIClientMixin, DetailView):
    """
    Vista detallada de un tipo de equipo
    """
    template_name = 'frontend/catalog/equipment_type_detail.html'
    context_object_name = 'equipment_type'
    login_url = 'frontend:login'
    
    def get_object(self, queryset=None):
        """Obtener el objeto desde la API"""
        try:
            api_client = self.get_api_client()
            type_id = self.kwargs.get('pk')
            
            response = api_client.get_equipment_type(type_id)
            
            if response:
                return response
            else:
                raise Http404("Tipo de equipo no encontrado")
                
        except APIException as e:
            logger.error(f"API error loading equipment type {self.kwargs.get('pk')}: {str(e)}")
            if e.status_code == 404:
                raise Http404("Tipo de equipo no encontrado")
            raise Http404("Error al cargar el tipo de equipo")
        except Exception as e:
            logger.error(f"Unexpected error loading equipment type {self.kwargs.get('pk')}: {str(e)}")
            raise Http404("Error al cargar el tipo de equipo")
    
    def get_context_data(self, **kwargs):
        """Agregar contexto adicional"""
        context = super().get_context_data(**kwargs)
        obj = self.get_object()
        
        context['title'] = f'Detalle: {obj.get("name", "")}'
        context['can_edit'] = True  # Aquí se pueden agregar permisos específicos
        context['can_delete'] = True  # Aquí se pueden agregar permisos específicos
        
        # URLs para acciones
        context['edit_url'] = reverse('frontend:equipment_type_edit', kwargs={'pk': obj.get('type_id')})
        context['delete_url'] = reverse('frontend:equipment_type_delete', kwargs={'pk': obj.get('type_id')})
        context['list_url'] = reverse('frontend:equipment_type_list')
        
        return context


class EquipmentTypeDeleteView(LoginRequiredMixin, APIClientMixin, DeleteView):
    """
    Vista para eliminar tipos de equipo con verificación de dependencias
    """
    template_name = 'frontend/catalog/equipment_type_confirm_delete.html'
    success_url = reverse_lazy('frontend:equipment_type_list')
    login_url = 'frontend:login'
    
    def get_object(self, queryset=None):
        """Obtener el objeto desde la API"""
        try:
            api_client = self.get_api_client()
            type_id = self.kwargs.get('pk')
            
            response = api_client.get_equipment_type(type_id)
            
            if response:
                return response
            else:
                raise Http404("Tipo de equipo no encontrado")
                
        except APIException as e:
            logger.error(f"API error loading equipment type {self.kwargs.get('pk')}: {str(e)}")
            if e.status_code == 404:
                raise Http404("Tipo de equipo no encontrado")
            raise Http404("Error al cargar el tipo de equipo")
        except Exception as e:
            logger.error(f"Unexpected error loading equipment type {self.kwargs.get('pk')}: {str(e)}")
            raise Http404("Error al cargar el tipo de equipo")
    
    def get_context_data(self, **kwargs):
        """Agregar contexto para confirmación de eliminación"""
        context = super().get_context_data(**kwargs)
        obj = self.get_object()
        
        context['title'] = f'Eliminar Tipo de Equipo: {obj.get("name", "")}'
        context['object'] = obj
        context['cancel_url'] = reverse('frontend:equipment_type_list')
        
        # Verificar dependencias (esto se puede expandir según las relaciones)
        context['has_dependencies'] = False
        context['dependencies'] = []
        
        # Aquí se pueden agregar verificaciones de dependencias
        # Por ejemplo, verificar si hay equipos usando este tipo
        
        return context
    
    def delete(self, request, *args, **kwargs):
        """Procesar eliminación"""
        try:
            api_client = self.get_api_client()
            type_id = self.kwargs.get('pk')
            obj = self.get_object()
            
            # Verificar dependencias antes de eliminar
            # (Esto se puede implementar con llamadas adicionales a la API)
            
            # Intentar eliminar
            response = api_client.delete_equipment_type(type_id)
            
            if response is not False:  # La API puede retornar None para DELETE exitoso
                messages.success(
                    request, 
                    f"Tipo de equipo '{obj.get('name', '')}' eliminado exitosamente."
                )
                return redirect(self.success_url)
            else:
                messages.error(
                    request, 
                    "No se puede eliminar el tipo de equipo. Puede estar en uso."
                )
                return redirect('frontend:equipment_type_detail', pk=type_id)
                
        except APIException as e:
            logger.error(f"API error deleting equipment type {self.kwargs.get('pk')}: {str(e)}")
            self.handle_api_error(e, "Error al eliminar el tipo de equipo.")
            return redirect('frontend:equipment_type_detail', pk=self.kwargs.get('pk'))
        except Exception as e:
            logger.error(f"Unexpected error deleting equipment type {self.kwargs.get('pk')}: {str(e)}")
            messages.error(
                request, 
                "Error al eliminar el tipo de equipo. Intenta nuevamente."
            )
            return redirect('frontend:equipment_type_detail', pk=self.kwargs.get('pk'))


# =============================================================================
# AJAX Views para funcionalidad dinámica
# =============================================================================

@method_decorator(csrf_exempt, name='dispatch')
class EquipmentTypeExportXLSView(LoginRequiredMixin, APIClientMixin, View):
    """
    Vista para exportar tipos de equipo a Excel
    """
    
    def get(self, request):
        """Exportar todos los tipos de equipo a Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from io import BytesIO
            
            # Obtener todos los tipos de equipo sin paginación
            api_client = self.get_api_client()
            
            # Obtener todos los registros
            all_equipment_types = []
            page = 1
            
            while True:
                response = api_client.get_equipment_types(page=page, page_size=100)
                if response and 'results' in response:
                    all_equipment_types.extend(response['results'])
                    if not response.get('next'):
                        break
                    page += 1
                else:
                    break
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tipos de Equipo"
            
            # Encabezados
            headers = [
                'ID', 'Código', 'Nombre', 'Categoría', 'Descripción', 
                'Activo', 'Icono', 'Fecha Creación', 'Fecha Actualización'
            ]
            
            # Escribir encabezados
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Escribir datos
            for row, eq_type in enumerate(all_equipment_types, 2):
                ws.cell(row=row, column=1, value=eq_type.get('type_id', ''))
                ws.cell(row=row, column=2, value=eq_type.get('type_code', ''))
                ws.cell(row=row, column=3, value=eq_type.get('name', ''))
                ws.cell(row=row, column=4, value=eq_type.get('category', ''))
                ws.cell(row=row, column=5, value=eq_type.get('description', ''))
                ws.cell(row=row, column=6, value='Sí' if eq_type.get('is_active', False) else 'No')
                ws.cell(row=row, column=7, value=eq_type.get('icon', ''))
                ws.cell(row=row, column=8, value=str(eq_type.get('created_at', ''))[:19] if eq_type.get('created_at') else '')
                ws.cell(row=row, column=9, value=str(eq_type.get('updated_at', ''))[:19] if eq_type.get('updated_at') else '')
            
            # Ajustar anchos de columna
            column_widths = [8, 15, 30, 15, 40, 8, 20, 20, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
            # Crear respuesta HTTP
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="tipos_equipo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            messages.success(request, f'Se exportaron {len(all_equipment_types)} tipos de equipo exitosamente.')
            return response
            
        except ImportError:
            messages.error(request, "La librería openpyxl no está instalada. Por favor contacte al administrador.")
            return redirect('frontend:equipment_type_list')
        except Exception as e:
            logger.error(f"Error exporting equipment types to XLS: {str(e)}")
            messages.error(request, "Error al exportar los tipos de equipo.")
            return redirect('frontend:equipment_type_list')


class EquipmentTypeExportPDFView(LoginRequiredMixin, APIClientMixin, View):
    """
    Vista para exportar tipos de equipo a PDF
    """
    
    def get(self, request):
        """Exportar todos los tipos de equipo a PDF"""
        try:
            from django.template.loader import render_to_string
            from io import BytesIO
            
            # Obtener todos los tipos de equipo
            api_client = self.get_api_client()
            
            all_equipment_types = []
            page = 1
            
            while True:
                response = api_client.get_equipment_types(page=page, page_size=100)
                if response and 'results' in response:
                    all_equipment_types.extend(response['results'])
                    if not response.get('next'):
                        break
                    page += 1
                else:
                    break
            
            # Preparar datos para el template
            context = {
                'equipment_types': all_equipment_types,
                'export_date': timezone.now(),
                'total_count': len(all_equipment_types)
            }
            
            # Renderizar template HTML para PDF
            html_string = render_to_string('frontend/catalog/equipment_type_export_pdf.html', context)
            
            # Generar PDF
            try:
                import weasyprint
                pdf_file = BytesIO()
                weasyprint.HTML(string=html_string).write_pdf(pdf_file)
                pdf_file.seek(0)
                
                response = HttpResponse(pdf_file, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="tipos_equipo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
                
                messages.success(request, f'Se exportaron {len(all_equipment_types)} tipos de equipo exitosamente.')
                return response
                
            except ImportError:
                # Si WeasyPrint no está disponible, devolver HTML
                response = HttpResponse(html_string, content_type='text/html')
                response['Content-Disposition'] = f'attachment; filename="tipos_equipo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.html"'
                messages.info(request, "Exportación en formato HTML (WeasyPrint no disponible)")
                return response
                
        except Exception as e:
            logger.error(f"Error exporting equipment types to PDF: {str(e)}")
            messages.error(request, "Error al exportar los tipos de equipo.")
            return redirect('frontend:equipment_type_list')


class EquipmentTypeAjaxSearchView(LoginRequiredMixin, APIClientMixin, View):
    """
    Vista AJAX para búsqueda dinámica de tipos de equipo
    """
    
    def get(self, request):
        """Búsqueda AJAX"""
        try:
            api_client = self.get_api_client()
            
            # Parámetros de búsqueda
            search = request.GET.get('q', '').strip()
            category = request.GET.get('category', '').strip()
            limit = int(request.GET.get('limit', 10))
            
            params = {
                'search': search,
                'page_size': limit
            }
            
            if category:
                params['category'] = category
            
            # Llamada a la API
            response = api_client.get_equipment_types(**params)
            
            if response and 'results' in response:
                results = []
                for item in response['results']:
                    results.append({
                        'id': item['type_id'],
                        'text': f"{item['name']} ({item['type_code']})",
                        'category': item['category'],
                        'is_active': item['is_active']
                    })
                
                return JsonResponse({
                    'success': True,
                    'results': results,
                    'total': response.get('count', 0)
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Error al buscar tipos de equipo'
                })
                
        except APIException as e:
            logger.error(f"API error in AJAX search: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': 'Error de conexión con el servidor'
            })
        except Exception as e:
            logger.error(f"Unexpected error in AJAX search: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': 'Error de conexión con el servidor'
            })


def equipment_type_check_code(request):
    """
    Vista AJAX para verificar unicidad de código
    """
    if request.method == 'GET':
        try:
            api_client = ForgeAPIClient(request=request)
            code = request.GET.get('code', '').strip().upper()
            exclude_id = request.GET.get('exclude_id', '')
            
            if not code:
                return JsonResponse({'available': True})
            
            # Buscar código existente
            params = {'search': code}
            response = api_client.get_equipment_types(**params)
            
            if response and 'results' in response:
                # Verificar si existe el código exacto
                for item in response['results']:
                    if item['type_code'] == code:
                        # Si estamos editando, excluir el ID actual
                        if exclude_id and str(item['type_id']) == str(exclude_id):
                            continue
                        return JsonResponse({
                            'available': False,
                            'message': f'El código {code} ya está en uso'
                        })
                
                return JsonResponse({'available': True})
            else:
                return JsonResponse({'available': True})
                
        except Exception as e:
            logger.error(f"Error checking code availability: {str(e)}")
            return JsonResponse({
                'available': True,  # En caso de error, permitir continuar
                'error': 'Error al verificar disponibilidad'
            })
    
    return JsonResponse({'available': True})


class EquipmentTypeSearchForm(forms.Form):
    """
    Formulario para búsqueda y filtrado de tipos de equipo
    """
    
    search = forms.CharField(
        max_length=100,
        required=False,
        widget=forms.TextInput(attrs={
            'class': 'form-control',
            'placeholder': 'Buscar por código, nombre o descripción...',
            'autocomplete': 'off'
        }),
        label='Búsqueda'
    )
    
    category = forms.ChoiceField(
        required=False,
        widget=forms.Select(attrs={
            'class': 'form-select'
        }),
        label='Categoría'
    )
    
    is_active = forms.ChoiceField(
        choices=[
            ('', 'Todos los estados'),
            ('true', 'Solo activos'),
            ('false', 'Solo inactivos')
        ],
        required=False,
        widget=forms.Select(attrs={
            'class': 'form-select'
        }),
        label='Estado'
    )
    
    def __init__(self, *args, **kwargs):
        # Extraer las categorías dinámicas si se proporcionan
        categories_choices = kwargs.pop('categories_choices', None)
        super().__init__(*args, **kwargs)
        
        # Si se proporcionaron categorías dinámicas, usarlas
        if categories_choices:
            self.fields['category'].choices = [('', 'Todas las categorías')] + categories_choices
        else:
            # Fallback a categorías reales de la BD
            CATEGORY_CHOICES = [
                ('AUTOMOTRIZ', 'Automotriz'),
                ('INDUSTRIAL', 'Industrial'),
                ('AGRICOLA', 'Agrícola'),
                ('CONSTRUCCION', 'Construcción'),
                ('ELECTRONICO', 'Electrónico'),
                ('CAT001', 'CAT TRX'),
                ('OTRO', 'Otro'),
            ]
            self.fields['category'].choices = [('', 'Todas las categorías')] + CATEGORY_CHOICES


# === VISTAS DE EXPORTACIÓN ===

class EquipmentTypeExportXLSView(LoginRequiredMixin, APIClientMixin, View):
    """
    Vista para exportar tipos de equipo a Excel
    """
    
    def get(self, request):
        """Exportar todos los tipos de equipo a Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from io import BytesIO
            
            # Obtener todos los tipos de equipo sin paginación
            api_client = self.get_api_client()
            
            # Obtener todos los registros
            all_equipment_types = []
            page = 1
            
            while True:
                response = api_client.get_equipment_types(page=page, page_size=100)
                if response and 'results' in response:
                    all_equipment_types.extend(response['results'])
                    if not response.get('next'):
                        break
                    page += 1
                else:
                    break
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tipos de Equipo"
            
            # Encabezados
            headers = [
                'ID', 'Código', 'Nombre', 'Categoría', 'Descripción', 
                'Activo', 'Icono', 'Fecha Creación', 'Fecha Actualización'
            ]
            
            # Escribir encabezados
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Escribir datos
            for row, eq_type in enumerate(all_equipment_types, 2):
                ws.cell(row=row, column=1, value=eq_type.get('type_id', ''))
                ws.cell(row=row, column=2, value=eq_type.get('type_code', ''))
                ws.cell(row=row, column=3, value=eq_type.get('name', ''))
                ws.cell(row=row, column=4, value=eq_type.get('category', ''))
                ws.cell(row=row, column=5, value=eq_type.get('description', ''))
                ws.cell(row=row, column=6, value='Sí' if eq_type.get('is_active', False) else 'No')
                ws.cell(row=row, column=7, value=eq_type.get('icon', ''))
                ws.cell(row=row, column=8, value=str(eq_type.get('created_at', ''))[:19] if eq_type.get('created_at') else '')
                ws.cell(row=row, column=9, value=str(eq_type.get('updated_at', ''))[:19] if eq_type.get('updated_at') else '')
            
            # Ajustar anchos de columna
            column_widths = [8, 15, 30, 15, 40, 8, 20, 20, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
            # Crear respuesta HTTP
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="tipos_equipo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            messages.success(request, f'Se exportaron {len(all_equipment_types)} tipos de equipo exitosamente.')
            return response
            
        except ImportError:
            messages.error(request, "La librería openpyxl no está instalada. Por favor contacte al administrador.")
            return redirect('frontend:equipment_type_list')
        except Exception as e:
            logger.error(f"Error exporting equipment types to XLS: {str(e)}")
            messages.error(request, "Error al exportar los tipos de equipo.")
            return redirect('frontend:equipment_type_list')


class EquipmentTypeExportPDFView(LoginRequiredMixin, APIClientMixin, View):
    """
    Vista para exportar tipos de equipo a PDF
    """
    
    def get(self, request):
        """Exportar todos los tipos de equipo a PDF"""
        try:
            from django.template.loader import render_to_string
            from io import BytesIO
            
            # Obtener todos los tipos de equipo
            api_client = self.get_api_client()
            
            all_equipment_types = []
            page = 1
            
            while True:
                response = api_client.get_equipment_types(page=page, page_size=100)
                if response and 'results' in response:
                    all_equipment_types.extend(response['results'])
                    if not response.get('next'):
                        break
                    page += 1
                else:
                    break
            
            # Preparar datos para el template
            context = {
                'equipment_types': all_equipment_types,
                'export_date': timezone.now(),
                'total_count': len(all_equipment_types)
            }
            
            # Renderizar template HTML para PDF
            html_string = render_to_string('frontend/catalog/equipment_type_export_pdf.html', context)
            
            # Generar PDF
            try:
                import weasyprint
                pdf_file = BytesIO()
                weasyprint.HTML(string=html_string).write_pdf(pdf_file)
                pdf_file.seek(0)
                
                response = HttpResponse(pdf_file, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="tipos_equipo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
                
                messages.success(request, f'Se exportaron {len(all_equipment_types)} tipos de equipo exitosamente.')
                return response
                
            except ImportError:
                # Si WeasyPrint no está disponible, devolver HTML
                response = HttpResponse(html_string, content_type='text/html')
                response['Content-Disposition'] = f'attachment; filename="tipos_equipo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.html"'
                messages.info(request, "Exportación en formato HTML (WeasyPrint no disponible)")
                return response
                
        except Exception as e:
            logger.error(f"Error exporting equipment types to PDF: {str(e)}")
            messages.error(request, "Error al exportar los tipos de equipo.")
            return redirect('frontend:equipment_type_list')
